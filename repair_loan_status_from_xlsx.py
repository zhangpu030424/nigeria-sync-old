#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从审计 Excel 修复目标库 loan.status（status ≠ 实际 / if=1）。

Excel Sheet1 列：loan_no, status, 实际, if …
  - if=1 或 (实际 is not None and status != 实际) → 需修
  - 按 loan_no 更新（表内 application_no 可能被截断，不可靠）

Usage:
  # 只出 plan + SQL，不写库
  python3 repair_loan_status_from_xlsx.py \\
    --xlsx ~/Downloads/Untitled.xlsx \\
    --plan-file /tmp/repair_loan_status_xlsx_plan.jsonl \\
    --sql-out /tmp/repair_loan_status_xlsx.sql

  # 写库（pymysql）
  python3 repair_loan_status_from_xlsx.py \\
    --env ./ng_migration.env --apply \\
    --xlsx ~/Downloads/Untitled.xlsx \\
    --plan-file /tmp/repair_loan_status_xlsx_plan.jsonl \\
    --workers 8 --batch 200

  # 已有 plan，只 apply
  python3 repair_loan_status_from_xlsx.py \\
    --env ./ng_migration.env --apply-only \\
    --plan-file /tmp/repair_loan_status_xlsx_plan.jsonl
"""
from __future__ import print_function

import argparse
import json
import sys
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pymysql
from pymysql.cursors import DictCursor

HERE = Path(__file__).resolve().parent


def load_env(path: Path) -> Dict[str, str]:
    cfg = {}  # type: Dict[str, str]
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        cfg[k.strip()] = v.strip().strip("'\"")
    return cfg


def connect_target(cfg: Dict[str, str]):
    return pymysql.connect(
        host=cfg["TARGET_HOST"],
        port=int(cfg.get("TARGET_PORT", "3306")),
        user=cfg["TARGET_USER"],
        password=cfg["TARGET_PASSWORD"],
        database=cfg.get("TARGET_DB", "ng"),
        charset="utf8mb4",
        cursorclass=DictCursor,
        connect_timeout=10,
        read_timeout=120,
        write_timeout=120,
        autocommit=False,
    )


def _to_int(val: Any) -> Optional[int]:
    if val is None or val == "":
        return None
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def load_fix_rows_from_xlsx(xlsx: Path, sheet: str = "Sheet1") -> List[dict]:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("需要 openpyxl：pip install openpyxl")

    wb = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit("sheet 不存在: {0} / {1}".format(sheet, wb.sheetnames))
    ws = wb[sheet]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    col = {str(h): i for i, h in enumerate(header) if h is not None}
    for need in ("loan_no", "status", "实际"):
        if need not in col:
            raise SystemExit("缺少列 {0}，实际列={1}".format(need, list(col)))

    out = []  # type: List[dict]
    for r in rows_iter:
        loan_no = str(r[col["loan_no"]] or "").strip()
        if not loan_no:
            continue
        old_st = _to_int(r[col["status"]])
        new_st = _to_int(r[col["实际"]])
        iff = r[col["if"]] if "if" in col else None
        if new_st is None:
            continue
        if old_st == new_st and iff != 1:
            continue
        if old_st == new_st:
            continue
        item = {
            "loan_no": loan_no,
            "old_status": old_st,
            "new_status": new_st,
            "sn": str(r[col["sn"]]).strip() if "sn" in col and r[col["sn"]] is not None else "",
            "application_no": (
                str(r[col["application_no"]]).strip()
                if "application_no" in col and r[col["application_no"]] is not None
                else ""
            ),
        }
        out.append(item)
    return out


def write_plan(path: Path, rows: List[dict]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fp:
        for row in rows:
            fp.write(json.dumps(row, ensure_ascii=False) + "\n")
    return len(rows)


def read_plan(path: Path) -> List[dict]:
    out = []  # type: List[dict]
    with path.open(encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if line:
                out.append(json.loads(line))
    return out


def write_sql(path: Path, rows: List[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fp:
        fp.write("-- repair loan.status from xlsx, rows={0}\n".format(len(rows)))
        fp.write("SET NAMES utf8mb4;\n")
        for row in rows:
            loan_no = str(row["loan_no"]).replace("'", "''")
            new_st = int(row["new_status"])
            old_st = row.get("old_status")
            if old_st is None:
                fp.write(
                    "UPDATE loan SET status={0} WHERE loan_no='{1}';\n".format(
                        new_st, loan_no,
                    )
                )
            else:
                fp.write(
                    "UPDATE loan SET status={0} WHERE loan_no='{1}' AND status={2};\n".format(
                        new_st, loan_no, int(old_st),
                    )
                )


def _apply_batch(spec: dict) -> Dict[str, int]:
    cfg = spec["cfg"]
    batch = spec["batch"]  # type: List[dict]
    conn = connect_target(cfg)
    ok = 0
    miss = 0
    err = 0
    try:
        with conn.cursor() as cur:
            for row in batch:
                loan_no = row["loan_no"]
                new_st = int(row["new_status"])
                old_st = row.get("old_status")
                if old_st is None:
                    cur.execute(
                        "UPDATE loan SET status=%s WHERE loan_no=%s",
                        (new_st, loan_no),
                    )
                else:
                    cur.execute(
                        "UPDATE loan SET status=%s WHERE loan_no=%s AND status=%s",
                        (new_st, loan_no, int(old_st)),
                    )
                if cur.rowcount and cur.rowcount > 0:
                    ok += 1
                else:
                    miss += 1
        conn.commit()
    except Exception as exc:
        err = len(batch)
        try:
            conn.rollback()
        except Exception:
            pass
        print("batch error: {0}".format(exc), flush=True)
    finally:
        conn.close()
    return {"ok": ok, "miss": miss, "err": err}


def apply_plan(
    cfg: Dict[str, str],
    plan: List[dict],
    workers: int,
    batch_size: int,
) -> Dict[str, int]:
    batches = []  # type: List[List[dict]]
    for i in range(0, len(plan), max(1, batch_size)):
        batches.append(plan[i:i + batch_size])
    stats = {"ok": 0, "miss": 0, "err": 0}
    workers = max(1, min(int(workers or 1), 16))
    print(
        "apply start rows={0} batches={1} workers={2}".format(
            len(plan), len(batches), workers,
        ),
        flush=True,
    )
    t0 = time.time()
    if workers == 1:
        for i, batch in enumerate(batches, 1):
            part = _apply_batch({"cfg": cfg, "batch": batch})
            for k in stats:
                stats[k] += part[k]
            if i % 10 == 0 or i == len(batches):
                print(
                    "apply progress {0}/{1} stats={2} elapsed={3:.1f}s".format(
                        i, len(batches), stats, time.time() - t0,
                    ),
                    flush=True,
                )
    else:
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futs = [
                pool.submit(_apply_batch, {"cfg": cfg, "batch": b})
                for b in batches
            ]
            done = 0
            for fut in as_completed(futs):
                part = fut.result()
                for k in stats:
                    stats[k] += part[k]
                done += 1
                if done % 10 == 0 or done == len(batches):
                    print(
                        "apply progress {0}/{1} stats={2} elapsed={3:.1f}s".format(
                            done, len(batches), stats, time.time() - t0,
                        ),
                        flush=True,
                    )
    print("apply done stats={0} elapsed={1:.1f}s".format(stats, time.time() - t0), flush=True)
    return stats


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Repair loan.status from audit xlsx")
    p.add_argument("--env", default=str(HERE / "ng_migration.env"))
    p.add_argument("--xlsx", default="")
    p.add_argument("--sheet", default="Sheet1")
    p.add_argument("--plan-file", default="/tmp/repair_loan_status_xlsx_plan.jsonl")
    p.add_argument("--sql-out", default="")
    p.add_argument("--apply", action="store_true", help="读 xlsx → plan → 写库")
    p.add_argument("--apply-only", action="store_true", help="只读 plan 写库")
    p.add_argument("--workers", type=int, default=8)
    p.add_argument("--batch", type=int, default=200)
    args = p.parse_args(argv)

    plan_path = Path(args.plan_file)

    if not args.apply_only:
        if not args.xlsx:
            print("需要 --xlsx", file=sys.stderr)
            return 2
        xlsx = Path(args.xlsx).expanduser().resolve()
        if not xlsx.is_file():
            print("xlsx not found: {0}".format(xlsx), file=sys.stderr)
            return 1
        rows = load_fix_rows_from_xlsx(xlsx, args.sheet)
        trans = Counter((r["old_status"], r["new_status"]) for r in rows)
        print("loaded fix rows={0}".format(len(rows)), flush=True)
        print("transitions={0}".format(dict(trans)), flush=True)
        n = write_plan(plan_path, rows)
        print("plan written {0} rows={1}".format(plan_path, n), flush=True)
        if args.sql_out:
            sql_path = Path(args.sql_out)
            write_sql(sql_path, rows)
            print("sql written {0}".format(sql_path), flush=True)
    else:
        if not plan_path.is_file():
            print("plan missing: {0}".format(plan_path), file=sys.stderr)
            return 1
        rows = read_plan(plan_path)
        print("loaded plan rows={0}".format(len(rows)), flush=True)

    if not args.apply and not args.apply_only:
        print("dry-run only（未写库）。加 --apply / --apply-only 才更新。", flush=True)
        return 0

    env_path = Path(args.env)
    if not env_path.is_file():
        print("env not found: {0}".format(env_path), file=sys.stderr)
        return 1
    cfg = load_env(env_path)
    apply_plan(cfg, rows, args.workers, args.batch)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
